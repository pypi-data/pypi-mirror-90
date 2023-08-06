# -*- coding: utf-8 -*-
"""
Created on Fri Jun 10 21:00:21 2016

@author: MaxBohnet
"""

import tempfile
import logging
import os
from argparse import ArgumentParser
from typing import List, Dict
import sqlite3 as db
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import sys
if (sys.platform == 'win32'
    and sys.version_info.major == 3
    and sys.version_info.minor >= 6):
    sys._enablelegacywindowsfsencoding()

import orca
from wiver.wiver_python import WIVER
from cythoninstallhelpers.configure_logger import SimLogger


@orca.injectable()
def project_folder() -> str:
    """ The Project folder (%TEMP% by default)

    """
    folder = tempfile.gettempdir()
    return folder


@orca.injectable()
def params_file(project_folder:str) -> str:
    """ The params-file

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    params_file : str
    """
    fn = 'params'
    file_path = os.path.join(project_folder, '{}.h5'.format(fn))
    return file_path


@orca.injectable()
def result_folder(project_folder:str) -> str:
    """The result folder

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    result_folder : str
"""
    dirname = 'matrices'
    file_path = os.path.join(project_folder, dirname)
    return file_path


@orca.injectable()
def matrix_file(project_folder:str) -> str:
    """ The matrix-file

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    matrix_file : str
    """
    fn = 'matrices'
    file_path = os.path.join(project_folder, '{}.h5'.format(fn))
    return file_path


@orca.injectable()
def zones_file(project_folder:str) -> str:
    """ The zonal_data-file

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    zones_file : str
    """
    fn = 'zonal_data'
    file_path = os.path.join(project_folder, '{}.h5'.format(fn))
    return file_path


@orca.injectable()
def result_file(project_folder: str) -> str:
    """
    The results-file
    Must not contain non-ascii-letters in the directory- or filename,
    because netcdf4-files fail otherwise

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    result_file : str
    """
    fn = 'results'
    file_path = os.path.join(project_folder, '{}.h5'.format(fn))
    return file_path


@orca.injectable()
def balancing_file(project_folder:str) -> str:
    """ The balancing-file

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    balancing_file : str
    """
    fn = 'balancing'
    file_path = os.path.join(project_folder, '{}.h5'.format(fn))
    return file_path


@orca.injectable()
def starting_ending_trips_file(project_folder:str) -> str:
    """ The Excel-File for starting and ending trips

    Parameters
    ----------
    project_folder : str

    Returns
    -------
    starting_ending_trip_file : str
    """
    fn = 'starting_ending_trips.xlsx'
    file_path = os.path.join(project_folder, fn)
    return file_path


@orca.injectable()
def max_iterations() -> str:
    """ maximum number of iterations"""
    return 1


@orca.injectable()
def wiver_files(params_file:str,
                matrix_file:str,
                zones_file:str,
                balancing_file:str,
                result_file:str) -> Dict[str, str]:
    """
    Returns
    -------
    wiver_files : dict
        a dictionary with the file path to the input data files
    """
    files = {'params': params_file,
             'matrices': matrix_file,
             'zonal_data': zones_file,
             'balancing': balancing_file,
             'results': result_file}
    return files


@orca.injectable()
def n_threads() -> int:
    """
    The number of threads to use
    if not defined, use the maximum number of threads available,
    maximum number equals to the number of groups
    """


@orca.injectable()
def wiver(wiver_files: Dict[str, str], n_threads: int) -> WIVER:
    """
    Create a WIVER-instance with the given input data

    Parameters
    ----------
    wiver_files:
        dict with the input files
    n_threads:
        the maximum number of threads to use

    """
    wiver = WIVER.read_from_netcdf(wiver_files, n_threads=n_threads)
    return wiver


@orca.injectable()
def scenario() -> str:
    """The Scenario Name"""
    return 'wiver'


@orca.injectable()
def groups_to_calculate() -> List[int]:
    """List of group ids to calculate"""
    return []


@orca.injectable()
def connection(project_folder: str) -> db.Connection:
    """
    database connection to an sqlite-file named `wiver.db3` in the project_folder
    to write zonal data into

    Parameters
    ----------
    project_folder:
        the folder, where the sqlite-file is created
    """
    fn = os.path.join(project_folder, 'wiver.db3')
    if (sys.platform == 'win32'
        and sys.version_info.major == 3
        and sys.version_info.minor >= 7):
        fn = os.fsencode(fn)
    connection = db.connect(fn)
    return connection


@orca.injectable()
def reset_balancing() -> bool:
    """if True, balancing factors will be reset to 1 before calculating"""
    return False


@orca.step()
def add_logfile(project_folder: str, scenario: str):
    """
    add Logfile to logger

    Parameters
    ----------
    project_folder:
        filepath of the project folder
    scenario:
        the name of the scenario
    """
    logger = SimLogger()
    logger.add_packages(['wiver'])
    logfile = os.path.join(project_folder, 'log')
    os.makedirs(logfile, exist_ok=True)
    logger.configure(logfile, scenario=scenario)


@orca.step()
def close_logfile():
    """
    close logfiles
    """
    logging.shutdown()


@orca.step()
def save_input_data(wiver: WIVER, wiver_files: dict):
    """
    save the input data from the wiver-model
    as HDF5-files to the wiver_files

    Parameters
    ----------
    wiver:
        the Wiver-model
    wiver_files:
        dict with the filenames
    """
    datasets = ('params', 'zonal_data', 'matrices',
                'balancing')
    wiver.define_datasets()
    wiver.merge_datasets()
    for dataset_name in datasets:
        fn = wiver_files[dataset_name]
        wiver.save_data(dataset_name, fn)


@orca.step()
def run_wiver(wiver: WIVER, wiver_files: dict, max_iterations: int):
    """
    calculate wiver model

    Parameters
    ----------
    wiver:
        wiver-model
    wiver_files:
        dict with the input files
    max_iterations:
        Maximum number of iterations
    """
    wiver.calc_with_balancing(max_iterations=max_iterations)
    wiver.save_results(wiver_files)


@orca.step()
def run_wiver_for_selected_groups(wiver: WIVER,
                                  wiver_files: dict,
                                  max_iterations: int,
                                  reset_balancing: bool,
                                  groups_to_calculate: List[int]):
    """
    calculate wiver model for selected groups only

    Parameters
    ----------
    wiver:
        wiver-model
    wiver-files :
        dict
    max_iterations:
        int
    groups_to_calculate:
        list of group numbers, that should be calculated
    """
    if reset_balancing:
        wiver.init_array('balancing_factor_gj')
        wiver.init_array('trips_to_destination_gj')

    if groups_to_calculate:
        wiver.active_g[:] = np.in1d(wiver.groups, groups_to_calculate)
    wiver.calc_with_balancing(max_iterations=max_iterations)
    wiver.save_results(wiver_files)

@orca.step()
def save_results(wiver: WIVER, wiver_files: dict, result_folder: str):
    """
    save result matrices of wiver-model

    Parameters
    ----------
    wiver:
        wiver-model
    wiver-files :
        dict
    result_folder: str
        the folder to store the calculated matrices
    """
    fn = wiver_files['results']
    wiver.read_data('results', fn)
    wiver.save_results_to_visum(result_folder, visum_format='BK')


@orca.step()
def save_detailed_results(wiver: WIVER,
                          wiver_files: dict,
                          result_folder: str,
                          groups_to_calculate: List[int]):
    """
    save detailed result matrices of wiver-model

    Parameters
    ----------
    wiver:
        wiver-model
    wiver-files:
        dict
    result_folder: str
        the folder to store the calculated matrices
    groups_to_calculate:
        list of int
    """
    if groups_to_calculate:
        wiver.active_g[:] = np.in1d(wiver.groups, groups_to_calculate)
    fn = wiver_files['results']
    wiver.read_data('results', fn)
    wiver.save_detailed_results_to_visum(result_folder, visum_format='BK')

@orca.step()
def calc_starting_ending_trips(wiver: WIVER,
                               wiver_files: dict,
                               starting_ending_trips_file: str,
                               connection: db.Connection):
    """
    calculate starting and ending trips per zone

    Parameters
    ----------
    wiver:
        wiver-model
    wiver-files:
        dict
    starting_ending_trips_file:
        str
    connection:
        An open Database Connection
    """
    fn = wiver_files['results']
    wiver.read_data('results', fn)
    df = wiver.calc_starting_and_ending_trips()
    df.to_sql(name='wiver', con=connection, if_exists='replace')

    sheetname = 'data'
    # write to existing excel-file
    if os.path.isfile(starting_ending_trips_file):
        book = load_workbook(starting_ending_trips_file)
        with pd.ExcelWriter(starting_ending_trips_file, engine='openpyxl') as writer:
            writer.book = book
            #  overwrite the sheet if exists
            if sheetname in book.sheetnames:
                writer.book.remove(writer.book[sheetname])
            df.to_excel(writer, sheet_name=sheetname)
    # otherwise create a new xlsx-file
    else:
        df.to_excel(starting_ending_trips_file, sheet_name=sheetname)


if __name__ == '__main__':
    parser = ArgumentParser(description="Commercial Trip Model Wiver")
    parser.add_argument('-f', '--folder', dest='project_folder',
                        help='Project folder',
                        required=True)
    parser.add_argument('-m', '--result-folder', dest='result_folder',
                        help='Result Matrix folder',
                        required=True)
    parser.add_argument('-s', '--scenario', dest='scenario',
                        help='Scenario Name', default='Wiver',)
    parser.add_argument('-i', '--iterations', dest='max_iterations',
                        help='Maximum number of iterations', type=int,
                        default=5,)
    parser.add_argument('-g', '--groups', dest='groups_to_calculate',
                        help='groups to calculate', type=int, nargs='+')
    parser.add_argument('-c', '--calc-starting', dest='calc_starting',
                        action='store_true',
                        help='calculate starting and ending trips', )
    parser.add_argument('-d', '--save-detailed', dest='save_detailed',
                        action='store_true',
                        help='save detailed results', )
    parser.add_argument('-r', '--reset-balancing', dest='reset_balancing',
                        action='store_true',
                        help='reset balancing parameters to 0', )

    options = parser.parse_args()
    for key, value in options._get_kwargs():
        orca.add_injectable(key, value)

    steps = [
        'add_logfile',
        ]

    if options.groups_to_calculate:
        steps.append('run_wiver_for_selected_groups')
    else:
        steps.append('run_wiver')

    steps.append('save_results')

    if options.save_detailed:
        steps.append('save_detailed_results')

    if options.calc_starting:
        steps.append('calc_starting_ending_trips')

    orca.run(steps)
