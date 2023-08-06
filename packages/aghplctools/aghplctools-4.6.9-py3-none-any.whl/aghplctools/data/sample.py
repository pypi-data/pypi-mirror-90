"""Data types for interacting directly with .D samples (e.g. reprocessing, loading signals directly)"""
import re
import os
import datetime
import bisect
import openpyxl
import pathlib
import xml.etree.ElementTree
import numpy as np
from typing import List, Union, Tuple
from aston.tracefile.agilent_uv import AgilentCSDAD2, string_read
from aston.tracefile.agilent_ms import AgilentMS
from unithandler.base import UnitFloat

# todo general
#   - look through acaml files to see if anything else is worth parsing
#   - work out the difference between AgilentCSDAD 1 and 2 and make DADSpectrum class general
#   - dig for time unit in acaml file
#   - convert to subpackage

# regex for float
_float_pattern = '\d+(.\d+)?'
# regex to match signal patterns
_signal_re = re.compile(
    '(?P<name>.+),\s+'
    f'sig=(?P<wavelength>{_float_pattern}),(?P<width>{_float_pattern})\s+'
    f'ref=(?P<ref_string>(?P<ref>off|(?P<ref_wl>{_float_pattern}),(?P<ref_width>{_float_pattern})))',
    re.IGNORECASE
)
# separate regex to catch simple wavelength,bandwidth specification
_simple_re = re.compile(
    f'(?P<wavelength>{_float_pattern}),(?P<width>{_float_pattern})'
)
# regex for matching acaml file extensions (upper or lower case)
_acaml_re = re.compile(
    '\.acam(_|l)$',
    re.IGNORECASE,
)
# sequence acaml re
_seq_acaml_re = re.compile(
    f'^sequence{_acaml_re.pattern}',
    re.IGNORECASE,
)
_samp_acaml_re = re.compile(
    f'^sample{_acaml_re.pattern}',
    re.IGNORECASE,
)
# prefix regex to deal with irritating prefixed tree names
_prefix_re = re.compile('(?P<prefix>.+)ACAML')
# regex to extract timezone from Agilent date time string
_agilent_dt_re = re.compile('(?P<dt>.+)[-+].+')


def strptime_agilent_dt(dt_string: str) -> datetime.datetime:
    """
    Performs strptime on Agilent datetime string

    :param dt_string: agilent datetime strings
    :return: parsed datetime object
    """
    observed_formats = [
        '%Y-%m-%dT%H:%M:%S.%f',  # with ms value
        '%Y-%m-%dT%H:%M:%S',  # w/o ms value
    ]
    for fmt in observed_formats:
        try:
            match = _agilent_dt_re.match(dt_string)
            return datetime.datetime.strptime(
                match.group('dt'),
                fmt
            )
        except ValueError:
            continue
    raise ValueError(f'The provided datetime string "{dt_string}" does not match any expected formats')


def bisect_slice(array, minimum_value: float, maximum_value: float) -> Tuple[int, int]:
    """
    Finds the slice indicies for a minimum and maximum value in an array.

    :param array: array like bisectable (assumes sorted)
    :param minimum_value: minimum value
    :param maximum_value: maximum value
    :return: slice indicies
    """
    return (
        bisect.bisect_left(array, minimum_value),
        bisect.bisect_left(array, maximum_value) + 1,
    )


def check_or_locate_file(path: Union[str, pathlib.Path], file_name: str) -> pathlib.Path:
    """
    Checks whether the provided path points to the provided file name. If not, checks whether the path is a directory
    and searches for the file in the directory. If there are multiple occurences of the provided file name in a
    directory, the first is returned.

    :param path: path to search
    :param file_name: target file name
    :return: path to desired file
    """
    if isinstance(path, pathlib.Path) is False:
        path = pathlib.Path(path)
    if path.name.lower() == file_name.lower():
        return path
    elif path.is_dir():
        try:
            return next(path.glob(file_name))
        except StopIteration:
            raise FileNotFoundError(f'no file name matching "{file_name}" could be found in path {path}')
    raise FileNotFoundError(f'the path {path} does not point to a "{file_name}" file')


def retrieve_metadata_from_channel(path: Union[str, pathlib.Path]) -> dict:
    """
    Retrieves metadata from a .CH file

    :param path: path to read
    :return: returns a dictionary containing metadata from the channel
    """
    with open(path, 'rb') as f:
        # read in signal information string
        f.seek(0x1075)
        signal = string_read(f)
        # read chemstation version
        f.seek(0xEE3)
        cs_version = string_read(f)
    return {
        'signal_spec': signal,
        'cs_version': cs_version,
    }


class DADSignalInfo(object):
    # default unit for wavelength values
    DEFAULT_WAVELENGTH_UNIT = 'nm'
    DEFAULT_TIME_UNIT = 'min'

    def __init__(self,
                 wavelength: Union[float, UnitFloat],
                 bandwidth: Union[float, UnitFloat] = 1.,
                 reference: Union["DADSignalInfo", str] = None,
                 name: str = None,
                 ):
        """
        Class describing a DAD signal and its parameters

        :param wavelength: wavelength for the signal
        :param bandwidth: band width for the wavelength (signal is centered on the wavelength with this width)
        :param reference: reference information for the signal
        :param name: convenience name for the signal
        """
        # todo add baseline codes and peak flags (Concepts of ChemStation > Integration Events > Baseline Codes)
        # set protected values to enable functionality of subclass
        self._wavelength = None
        self._bandwidth = None
        if type(reference) is str:
            if reference.lower() == 'off':  # catch string equivalent to None
                reference = None
            else:
                reference = DADSignalInfo.create_from_agilent_string(
                    reference,
                    name_override='Ref',
                )
        self.reference = reference
        self.wavelength = wavelength
        self.bandwidth = bandwidth
        self.name = name

    @property
    def wavelength(self) -> UnitFloat:
        """Wavelength for the signal"""
        return self._wavelength

    @wavelength.setter
    def wavelength(self, value: Union[UnitFloat, float]):
        if type(value) is float:
            value = UnitFloat(
                value,
                self.DEFAULT_WAVELENGTH_UNIT,
            )
        self._wavelength = value

    @property
    def bandwidth(self) -> UnitFloat:
        """bandwidth for the signal band"""
        return self._bandwidth

    @bandwidth.setter
    def bandwidth(self, value: Union[UnitFloat, float]):
        if value is None:
            value = UnitFloat(
                1.,
                self.DEFAULT_WAVELENGTH_UNIT,
            )
        elif type(value) is float:
            value = UnitFloat(
                value,
                self.DEFAULT_WAVELENGTH_UNIT,
            )
        self._bandwidth = value

    @bandwidth.deleter
    def bandwidth(self):
        self.bandwidth = 1.

    @property
    def reference(self) -> "DADSignalInfo":
        """Reference band for the signal band"""
        return self._reference

    @reference.setter
    def reference(self, value: Union[str, "DADSignalInfo"]):
        if type(value) is str:
            value = self.create_from_agilent_string(value)
        self._reference = value

    @reference.deleter
    def reference(self):
        self._reference = None

    def __str__(self):
        return (
            f'{f"{self.name} " if self.name is not None else ""}'
            f'{self.wavelength} ({self.bandwidth})'
            f'{f" {self.reference}" if self.reference is not None else ""}'
        )

    def __repr__(self):
        return (
            f'{self.__class__.__name__}('
            f'{self.wavelength.real},'
            f'{self.bandwidth.real},'
            f'{self.reference},'  # todo adjust return of reference 
            f'{self.name}'
            f')'
        )

    @property
    def agilent_specification_string(self) -> str:
        """the specification string describing this instance (can be passed to create_from_string to reinstantiate)"""
        out = (
            f'{self.name}, '
            f'Sig={int(self.wavelength)},{int(self.bandwidth)}'
        )
        if self.reference is not None:
            out += f' Ref={int(self.reference.wavelength)},{int(self.reference.bandwidth)}'
        else:
            out += ' Ref=off'
        return out

    @classmethod
    def get_values_from_agilent_string(cls, string: str) -> dict:
        """
        Parses a standard Agilent signal description string (e.g. 'DAD1 A, Sig=210,4 Ref=360,100') and returns
        a dictionary of parsed values (can be used to instantiate a DADSignalInfo instance).

        :param string: signal description string
        :return: dictionary of parameters
        """
        match = _signal_re.match(string)
        if match is not None:
            return ({
                'wavelength': float(match.group('wavelength')),
                'bandwidth': float(match.group('width')),
                'reference': match.group('ref_string'),
                'name': match.group('name'),
            })
        else:
            # attempt to match a basic pattern
            match = _simple_re.match(string)
            if match is None:
                raise ValueError(f'The string "{string}" could not be interpreted as a DADSignal')
            return ({
                'wavelength': float(match.group('wavelength')),
                'bandwidth': float(match.group('width')),
            })

    @classmethod
    def create_from_agilent_string(cls,
                                   string: str,
                                   name_override: str = None,
                                   ) -> 'DADSignalInfo':
        """
        Creates a class instance from a standard Agilent signal description string (e.g. 'DAD1 A, Sig=210,4 Ref=360,100')

        :param string: signal description string
        :param name_override: override for name specification
        :return: DADSignal object
        """
        # parse string with class method
        parsed = cls.get_values_from_agilent_string(string)
        if name_override is not None:  # apply name override if specified
            parsed['name'] = name_override
        return cls(**parsed)

    @classmethod
    def create_from_CH_file(cls, file_path: Union[str, pathlib.Path]) -> "DADSignalInfo":
        """
        Creates a DADSignal info instance from a channel file.

        :param file_path: target file path
        """
        vals = retrieve_metadata_from_channel(file_path)
        return cls.create_from_agilent_string(vals['signal_spec'])

    @classmethod
    def get_signals_in_directory(cls, file_path: Union[str, pathlib.Path]) -> List["DADSignalInfo"]:
        """
        Creates a list of signals based on the .CH files in a directory.

        :param file_path: path to target directory
        :return: list of signal info objects
        """
        if isinstance(file_path, pathlib.Path) is False:
            file_path = pathlib.Path(file_path)
        if file_path.is_dir() is False:
            raise ValueError(f'the provided path is not a directory: {file_path}')
        out = []
        for ch_file in file_path.glob('*.CH'):
            out.append(cls.create_from_CH_file(ch_file))
        return out


class HPLCSampleInfo(object):
    def __init__(self,
                 sample_file_name: str,
                 method_name: str,
                 signals: Union[List[DADSignalInfo], List[str]],
                 datetimestamp: Union[str, datetime.datetime] = None,
                 ):
        """
        Data class for describing an HPLC sample.

        :param sample_file_name: name for sample
        :param datetimestamp: date and time stamp for when the sample was run
        :param method_name: name of method used to run the sample
        :param signals: list of signals associated with the run
        """
        self.sample_file_name = sample_file_name
        if type(datetimestamp) is str:
            datetimestamp = datetime.datetime.strptime(
                _agilent_dt_re.match(datetimestamp).group('dt'),
                '%Y-%m-%dT%H:%M:%S.%f'
            )
        self.datetimestamp: datetime.datetime = datetimestamp
        self.method_name = method_name
        # create signals from string if not DADSignalInfo
        if len(signals) > 0 and type(signals[0]) is str:
            parsed_signals = []
            for string in signals:
                try:
                    parsed_signals.append(DADSignalInfo.create_from_agilent_string(string))
                except ValueError:  # ignore non-DAD values
                    pass
            signals = parsed_signals
        self.signals: List[DADSignalInfo] = signals

    def __str__(self):
        return (
            f'{self.sample_file_name} run on {self.datetimestamp} with {len(self.signals)} signals'
        )

    def __repr__(self):
        return f'{self.__class__.__name__}({self.sample_file_name})'

    @property
    def date(self) -> str:
        """date which the sample was run on"""
        return str(self.datetimestamp.date())

    @property
    def timestamp(self) -> str:
        """Time of the day when the sample was run"""
        return str(self.datetimestamp.time())

    def as_dict(self) -> dict:
        """Returns the sample data as a dictionary"""
        return {
            'sample_file_name': self.sample_file_name,
            'datetimestamp': str(self.datetimestamp),
            'method_name': self.method_name,
            'signals': [str(signal) for signal in self.signals]
        }

    @staticmethod
    def find_acaml(acaml_path: Union[str, pathlib.Path]) -> xml.etree.ElementTree.ElementTree:
        """
        Finds an acaml file and loads the element tree

        :param acaml_path: path to acaml file or directory containing acaml file
        """
        if isinstance(acaml_path, pathlib.Path) is False:  # type convert to path if raw string
            acaml_path = pathlib.Path(acaml_path)
        if acaml_path.is_dir():  # not a acaml file
            seq_files = [filename for filename in os.listdir(acaml_path) if _seq_acaml_re.match(filename) is not None]
            samp_files = [filename for filename in os.listdir(acaml_path) if _samp_acaml_re.match(filename) is not None]
            if len(seq_files) == 1:
                acaml_path = acaml_path / seq_files[0]
            elif len(samp_files) == 1:
                acaml_path = acaml_path / samp_files[0]
            else:
                raise IOError(
                    f'{"Multiple" if len(seq_files) > 1 else "No"}'
                    f' sequence.acam_ or sample.acaml files were found in the directory "{acaml_path}". '
                    f'Please refine your search.'
                )
        else:
            if _acaml_re.search(acaml_path.suffix) is None:
                # todo consider trying to parse anyway
                raise ValueError(f'The file "{acaml_path}" does not have a valid extension')
        return xml.etree.ElementTree.parse(acaml_path)

    @classmethod
    def get_values_from_result_xml(cls, xml_path: Union[str, pathlib.Path]) -> dict:
        """
        Retrieves values from a Result.xml file. This is an old-style ChemStation
        metadata file (~B.04.03 era).

        :param xml_path: path to xml or directory containing xml file
        """
        xml_path = check_or_locate_file(xml_path, 'Result.xml')
        tree = xml.etree.ElementTree.parse(xml_path)
        root = tree.getroot()

        sample_info = [val for val in root.iter('SampleInformation')][0]
        sample_name = [val for val in sample_info.iter('SampleName')][0].text
        method_name = [val for val in sample_info.iter('Method')][0].text
        dt = datetime.datetime.strptime(
            [val for val in sample_info.iter('InjectionDateTime')][0].text,
            '%d-%b-%y, %H:%M:%S',
        )
        signals = []
        for signal in root.iter('Signal'):
            signals.append(signal.find('Description').text)

        return {
            'sample_file_name': sample_name,
            'method_name': method_name,
            'datetimestamp': dt,
            'signals': signals,
        }

    @classmethod
    def get_values_from_sample_xml(cls, xml_path: Union[str, pathlib.Path]) -> dict:
        """
        Retrieves values from a Sample.xml file. From ChemStation C.01.07

        :param xml_path: path to xml or directory containing xml file
        """
        xml_path = check_or_locate_file(xml_path, 'Sample.xml')
        tree = xml.etree.ElementTree.parse(xml_path)
        root = tree.getroot()
        sample_name = next(root.iter('Name')).text
        method_path = pathlib.Path(next(root.iter('ACQMethodPath')).text)
        signals = DADSignalInfo.get_signals_in_directory(
            xml_path.parent
        )
        return {
            'sample_file_name': sample_name,
            'method_name': method_path.name,
            'signals': signals,
        }

    @classmethod
    def get_values_from_xml(cls, xml_path: Union[str, pathlib.Path]) -> dict:
        """
        Attempts to find a Result.xml file and parse sample information from that.

        :param xml_path: path to xml or directory containing xml file
        """
        # todo make this not terrible
        try:
            return cls.get_values_from_sample_xml(xml_path)
        except FileNotFoundError:
            pass
        try:
            return cls.get_values_from_result_xml(xml_path)
        except FileNotFoundError:
            pass
        raise FileNotFoundError(f'no known sample metadata files could be found in the path {xml_path}')

    @classmethod
    def get_values_from_acaml(cls,
                              acaml: Union[
                                  str,
                                  pathlib.Path,
                                  xml.etree.ElementTree.ElementTree,
                              ]
                              ) -> dict:
        """
        Gets relevant values from an acaml file. (use sequence.acam_ in the desired .D folder)

        :param acaml: path to acaml file or parsed element tree root
        :return: dictionary of values of interest
        """
        # if provided with a path, create element tree
        if isinstance(acaml, xml.etree.ElementTree.ElementTree) is False:
            acaml = cls.find_acaml(acaml)

        root = acaml.getroot()
        prefix = _prefix_re.match(root.tag).group('prefix')

        # extract creation date
        cd = [val for val in root.iter(f'{prefix}CreationDate')][0]
        dt = strptime_agilent_dt(cd.text)

        # extract method name
        method = [val for val in root.iter(f'{prefix}Method')][0]
        method_name = [val.text for val in method.iter(f'{prefix}Name')][0]

        # sample name
        injection = [val for val in root.iter(f'{prefix}Injections')][0]
        # todo find better way to search this (multiple Name values in Injections tree)
        sample_name = [val.text for val in injection.iter(f'{prefix}Name')][0]

        signals = []
        for signal in root.iter(f'{prefix}Signal'):
            descrip = signal.find(f'{prefix}Description')
            try:
                signals.append(descrip.text)
            except ValueError:
                pass

        return {
            'sample_file_name': sample_name,
            'method_name': method_name,
            'datetimestamp': dt,
            'signals': signals,
        }

    @classmethod
    def find_and_get_metadata(cls, target_path: Union[str, pathlib.Path]) -> dict:
        """
        Attempts to locate and parse metadata files in both old (Result.xml) and new (ACAML) formats.
        If neither file type can be found, an error will be raised.

        :param target_path: target path to search
        :return: parsed dictionary for creating HPLCSampleInfo instance
        """
        try:
            return cls.get_values_from_acaml(target_path)
        except (IOError, ValueError):
            pass
        try:
            return cls.get_values_from_xml(target_path)
        except FileNotFoundError:
            pass
        raise FileNotFoundError(f'a metadata file could not be identified at or in the path {target_path}')

    @classmethod
    def create_from_acaml(cls, acaml: Union[str, xml.etree.ElementTree.ElementTree]) -> "HPLCSampleInfo":
        """
        Creates sample structure from an acaml file. (use sequence.acam_ in the desired .D folder)

        :param acaml: path to acaml file or parsed element tree root
        :return: parsed Sample instance
        """
        return cls(
            **cls.get_values_from_acaml(acaml)
        )

    @classmethod
    def create_from_xml(cls, xml_path: Union[str, xml.etree.ElementTree.ElementTree]) -> "HPLCSampleInfo":
        """
        Creates sample structure from a Sample.xml file (old style metadata) in the desired .D folder)

        :param xml_path: path to xml file or parsed element tree root
        :return: parsed Sample instance
        """
        return cls(
            **cls.get_values_from_xml(xml_path)
        )

    @classmethod
    def auto_create(cls, target_path: Union[str, pathlib.Path]) -> "HPLCSampleInfo":
        """
        Attempts to automatically create an instance from metadata in the target folder

        :param target_path: path to metadata file or folder containing metadata files
        :return: HPLCSampleInfo instance
        """
        return cls(
            **cls.find_and_get_metadata(target_path)
        )


class DADSpectrum(AgilentCSDAD2):
    def __init__(self,
                 filename=None,
                 ftype=None,
                 data=None,
                 ):
        """
        An object describing an Agilent DAD spectrum for a sample. Inherits Aston AgilentCSDAD2 and has additional
        methods for retrieving band information.

        :param filename: target filetype
        :param ftype:
        :param data:
        """
        # initialize object
        AgilentCSDAD2.__init__(self, filename=filename, ftype=ftype, data=data)
        # todo rewrite data attribute so that the data isn't extracted every time .data is accessed
        #   - also close the open file
        self.chromatogram = self.data

        # sort the data to enable indexing (data is stored from some middle value to max, then min to that middle value)
        max_index = self.chromatogram.columns.index(max(self.chromatogram.columns))
        reorder = [val for val in range(max_index + 1, len(self.chromatogram.columns))]
        reorder.extend([val for val in range(0, max_index + 1)])
        new_order = np.asarray(reorder)

        self.chromatogram.columns = np.asarray(self.chromatogram.columns)[new_order].tolist()
        self.chromatogram.values = self.chromatogram.values[:, new_order]

    def __str__(self):
        return f'{self.__class__.__name__} {self.info["name"]} {self.info["filename"].split(os.sep)[-1]}'

    @property
    def retention_times(self) -> np.ndarray:
        """retention times corresponding to the data array (min)"""
        return self.chromatogram.index

    @property
    def wavelengths(self) -> list:
        """list of wavelengths for the DAD"""
        # todo check return type
        return self.chromatogram.columns

    @property
    def total_absorbance_chromatogram(self) -> np.ndarray:
        """
        The total absorbance chromatogram for the spectrum
        (sum of all intensities for each retention time)
        """
        return self.chromatogram.values.sum(axis=1)

    @property
    def maximum_wavelength_array(self) -> np.ndarray:  # todo
        """Array of the wavelengths for the maximum intensity at each retention time"""
        return

    def _get_band_indicies(self, wavelength: float, bandwidth: float = 1.) -> Tuple[int, int]:
        """
        Retrieves the band indicies for the wavelength and band width described.

        :param wavelength: wavelength
        :param bandwidth: band width
        :return: tuple of indicies which encompass the band
        """
        # todo catch and raise errors if invalid indicies are generated
        # todo set up index tests to verify slicing operates as intended
        return bisect_slice(
            self.chromatogram.columns,
            wavelength - bandwidth / 2,
            wavelength + bandwidth / 2,
        )

    def _get_retention_indicies(self, start_time: float, end_time: float) -> Tuple[int, int]:
        """
        Gets the appropriate indicies to include the provided retention times in the retention time array

        :param start_time: start retention time (min)
        :param end_time: end retention time (min)
        """
        return bisect_slice(self.retention_times, start_time, end_time)

    def get_band_wavelengths(self, wavelength: float, bandwidth: float = 1.) -> list:
        """
        Returns a list of wavelengths corresponding to the band specified.

        :param wavelength: wavelength
        :param bandwidth: band width
        :return:
        """
        # get indicies for described slice
        ind_left, ind_right = self._get_band_indicies(wavelength=wavelength, bandwidth=bandwidth)
        return self.chromatogram.columns[ind_left:ind_right]

    def get_band_intensities(self, wavelength: float, bandwidth: float = 1.) -> np.ndarray:
        """
        Retrieve array of values described by the wavelength and band width described. The returned array will have shape
        [wavelength, retention time]. The corresponding wavelengths are given by DADSpectrum.get_band_wavelengths and
        the retention times by DADSpectrum.retention_times.

        :param wavelength: wavelength
        :param bandwidth: band width
        :return: array of band intensities
        """
        # todo account for resolution of recorded spectrum
        # get indicies for described slice
        ind_left, ind_right = self._get_band_indicies(wavelength=wavelength, bandwidth=bandwidth)
        return self.chromatogram.values[:, ind_left:ind_right]

    def get_band_mean_intensity(self, wavelength: float, bandwidth: float = 1.) -> np.ndarray:
        """
        Retrieve the intensity array described by the wavelength and bandwidth described. The returned array will be the
        mean of the intensities in the band (wavelength - bandwidth / 2, wavelength + bandwidth / 2).

        :param wavelength: wavelength
        :param bandwidth: band width
        :return: array of mean intensities
        """
        return self.get_band_intensities(
            wavelength=wavelength,
            bandwidth=bandwidth,
        ).mean(axis=1)

    def get_intensities_from_signal(self, signal: DADSignalInfo) -> np.ndarray:
        """
        Retrieve the intensity array described by the DADSignalInfo object.

        :param signal: signal descriptor
        :return: array of mean intensities
        """
        return self.get_band_intensities(
            signal.wavelength.real,
            signal.bandwidth.real,
        )

    def get_component_spectrum(self, retention_start: float, retention_end: float) -> np.ndarray:
        """
        Retrieves the component spectrum for the provided retention time slice.

        :param retention_start: retention time start
        :param retention_end: retention time end
        :return:
        """
        start_ind, end_ind = self._get_retention_indicies(retention_start, retention_end)
        return np.stack((
            self.wavelengths,
            self.chromatogram.values[start_ind:end_ind].sum(axis=0),
        ))

    @classmethod
    def create_from_D_file(cls, file_path: Union[pathlib.Path, str]) -> "DADSpectrum":
        """
        Creates a DADSpectrum instance from an Agilent .D file

        :param file_path: path to .D sample file
        :return: interpreted .D file with metadata and loaded UV data
        """
        if type(file_path) is str:
            file_path = pathlib.Path(file_path)
        if file_path.suffix != '.D' or os.path.isdir(file_path) is False:
            raise ValueError(f'The file path "{file_path}" does not appear to be a valid Agilent sample directory')
        uv_files = [filename for filename in os.listdir(file_path) if filename.endswith('.UV')]
        # todo catch if multiple UV files (does this happen?)
        try:
            return cls(
                os.path.join(file_path, uv_files[0])
            )
        except IndexError:
            raise FileNotFoundError(f'no .UV files were found in the target path')

    def write_to_allotrope(self, filename: str):  # todo
        raise NotImplementedError('there is no way to do this currently')


class MSSpectrum(AgilentMS):
    def __init__(self,
                 filename=None,
                 ftype=None,
                 data=None,
                 ):
        """
        An object describing an Agilent DAD spectrum for a sample. Inherits Aston AgilentCSDAD2 and has additional
        methods for retrieving band information.

        :param filename: target filetype
        :param ftype:
        :param data:
        """
        # initialize object
        AgilentMS.__init__(self, filename=filename, ftype=ftype, data=data)
        # todo rewrite data attribute so that the data isn't extracted every time .data is accessed
        self.chromatogram = self.data

        # generate sorting array
        sort_array = np.arange(self.chromatogram.columns.shape[0])[self.chromatogram.columns.argsort()]
        self.chromatogram.columns: np.ndarray = self.chromatogram.columns[sort_array]  # sort mass array
        # todo type hint for csr matrix
        self.chromatogram.values = self.chromatogram.values[:, sort_array]  # rearrange csr matrix for sorted mzs

    def __str__(self):
        return f'{self.__class__.__name__} {self.info["name"]} {self.info["filename"].split(os.sep)[-1]}'

    @property
    def retention_times(self) -> np.ndarray:
        """retention times corresponding to the data array (min)"""
        return self.chromatogram.index

    @property
    def masses(self) -> np.ndarray:
        """array of wavelengths for the DAD"""
        return self.chromatogram.columns

    @property
    def summed_intensity_array(self) -> np.ndarray:
        """returns the summed intensity array of the spectrum"""
        return self.chromatogram.values.toarray().sum(axis=0)

    @property
    def summed_spectrum(self) -> np.ndarray:
        """returns the mz and summed intensity array for the entire run"""
        return np.stack((self.masses, self.summed_intensity_array))

    def _get_mass_indicies(self, start_mz: float, end_mz: float = None) -> Tuple[int, int]:
        """
        Determines and returns the appropriate indicies for the provided masses in the mass array.

        :param start_mz: start m/z ratio for the region
        :param end_mz: end m/z ratio for the region. If not specified, the closest value to the start m/z will be
            returned
        """
        return (
            bisect.bisect_left(self.masses, start_mz),
            bisect.bisect_left(self.masses, end_mz or start_mz) + 1,
        )

    def _get_retention_indicies(self, start_time: float, end_time: float) -> Tuple[int, int]:
        """
        Gets the appropriate indicies to include the provided retention times in the retention time array

        :param start_time: start retention time (min)
        :param end_time: end retention time (min)
        """
        return bisect_slice(self.retention_times, start_time, end_time)

    def get_ion_intensities(self, start_mz: float, end_mz: float = None) -> np.ndarray:
        """
        Returns the intensity integral array (reconstructed single ion monitoring) for the provided ion m/z window.

        :param start_mz: start m/z ratio for the region
        :param end_mz: end m/z ratio for the region.
        """
        start_ind, end_ind = self._get_mass_indicies(start_mz, end_mz)
        return self.chromatogram.values.toarray()[:, start_ind:end_ind].sum(axis=1)

    def get_spectrum_of_retention_period(self, start_time: float, end_time: float) -> np.ndarray:
        """
        Returns the intensity array for the mass spectrum in the retention time region provided.

        :param start_time: start retention time (min)
        :param end_time: end retention time (min)
        """
        start_ind, end_ind = self._get_retention_indicies(start_time, end_time)
        return self.chromatogram.values.toarray()[start_ind:end_ind].sum(axis=0)

    @classmethod
    def create_from_D_file(cls, file_path: Union[pathlib.Path, str]) -> "MSSpectrum":
        """
        Creates a MSSpectrum instance from an Agilent .D file

        :param file_path: path to .D file
        :return: instance
        """
        if type(file_path) is str:
            file_path = pathlib.Path(file_path)
        if file_path.suffix != '.D' or os.path.isdir(file_path) is False:
            raise ValueError(f'The file path "{file_path}" does not appear to be a valid Agilent sample directory')
        ms_files = [filename for filename in os.listdir(file_path) if filename.endswith('.MS')]
        # todo catch if multiple MS files (does this happen?)
        try:
            return cls(
                os.path.join(file_path, ms_files[0])
            )
        except IndexError:
            raise FileNotFoundError(f'no .MS files were found in the target path')

    @staticmethod
    def _find_some_peaks(y, n_peaks: int):
        """roughly locates 4 peaks by maximum values in the spectrum and returns their index"""
        split = int(len(y) / n_peaks)
        start = 0
        end = start + split
        splity = []
        for i in range(n_peaks):
            splity.append(np.asarray(y[start:end]))
            start += split
            end += split
        out = []
        for ind, section in enumerate(splity):
            maxy = max(section)
            if maxy == max(section[1:-1]):  # if max is not at the edge of the spectrum
                out.append(np.where(section == maxy)[0][0] + split * ind)
        return out

    @staticmethod
    def _calculate_resolution(x: np.ndarray, y: np.ndarray, index=None, threshold=5):
        """
        Finds the resolution and full width at half max of a spectrum. From PythoMS.

        :param x: list of mz values
        :param y: corresponding list of intensity values
        :param index: index of maximum intensity (optional; used if the resolution of a specific peak is desired)
        :param threshold: signal to noise threshold required to output a resolution
        :return: resolution
        """
        y = np.asarray(y)  # convert to array for efficiency
        if index is None:  # find index and value of maximum
            maxy = max(y)
            index = np.where(y == maxy)[0][0]
        else:
            maxy = y[index]
        # if intensity to average is below this threshold (rough estimate of signal/noise)
        if maxy / (sum(y) / len(y)) < threshold:
            return None
        halfmax = maxy / 2
        indleft = int(index) - 1  # generate index counters for left and right walking
        indright = int(index) + 1
        while y[indleft] > halfmax:  # while intensity is still above halfmax
            indleft -= 1
        while y[indright] > halfmax:
            indright += 1
        return x[index] / (x[indright] - x[indleft])  # return resolution (mz over full width at half max)

    def auto_resolution(self, npeaks: int = 4) -> float:
        """
        Attempts to automatically determine the resolution of the spectrum.

        :param npeaks: number of peakds to try to find
        :return: estimated resolution
        """
        intensity_array = self.summed_intensity_array
        peak_indicies = self._find_some_peaks(
            intensity_array,
            n_peaks=npeaks,
        )
        resolutions = [
            self._calculate_resolution(
                self.masses,
                intensity_array,
                ind,
                threshold=10,
            )
            for ind in peak_indicies
        ]
        resolutions = [res for res in resolutions if res is not None]
        return sum(resolutions) / len(resolutions)

    def extract_function_time_tic(self):
        """duck-type method for PythoMS"""
        pass

    def get_tic_of_function(self, function: int) -> np.ndarray:
        """duck-type method for retrieving the TIC (expected in PythoMS)"""
        return self.chromatogram.values.toarray().sum(axis=1)

    def get_timepoints_of_function(self, function: int) -> np.ndarray:
        """duck-type method for retrieving the timepoints (expected in PythoMS)"""
        return self.retention_times

    @property
    def functions(self):
        """duck type function information (expected in PythoMS)"""
        return {
            1: {
                'sr': [1, self.retention_times.shape[0]],
                'nscans': self.retention_times.shape[0],
                'type': 'MS',
                'mode': '',
                'window': [self.masses.argmin(), self.masses.argmax()]
            }
        }


class DADSignal(DADSignalInfo):
    def __init__(self,
                 wavelength: Union[float, UnitFloat],
                 bandwidth: Union[float, UnitFloat] = 1.,
                 reference: Union["DADSignal", DADSignalInfo, str] = None,
                 name: str = None,
                 spectrum: DADSpectrum = None,
                 ):
        """
        Class describing a DAD signal and its data.

        :param wavelength: wavelength for the signal
        :param bandwidth: band width for the wavelength (signal is centered on the wavelength with this width)
        :param reference: reference information for the signal
        :param name: convenience name for the signal
        :param spectrum: a DADSpectrum object which will be referenced for retrieving data.
        """
        # todo create catch for wavelengths outside of spectrum
        self.spectrum: DADSpectrum = spectrum
        self._unreferenced_intensities: np.ndarray = None
        self._reference: DADSignal = None  # modify type hint
        DADSignalInfo.__init__(
            self,
            wavelength=wavelength,
            bandwidth=bandwidth,
            reference=reference,
            name=name,
        )
        self._update_unreferenced_intensities()

    @property
    def wavelength(self) -> UnitFloat:
        """wavelength for the signal"""
        return self._wavelength

    @wavelength.setter
    def wavelength(self, value):
        if isinstance(value, UnitFloat) is False:
            value = UnitFloat(
                value,
                self.DEFAULT_WAVELENGTH_UNIT,
            )
        self._wavelength = value
        self._update_unreferenced_intensities()

    @property
    def bandwidth(self) -> UnitFloat:
        """band width for the signal band"""
        return self._bandwidth

    @bandwidth.setter
    def bandwidth(self, value: Union[UnitFloat, float]):
        if value is None:  # if no value, set to None
            value = UnitFloat(
                1.,
                self.DEFAULT_WAVELENGTH_UNIT,
            )
        elif isinstance(value, UnitFloat) is False:
            value = UnitFloat(
                value,
                self.DEFAULT_WAVELENGTH_UNIT,
            )
        self._bandwidth = value
        self._update_unreferenced_intensities()

    @bandwidth.deleter
    def bandwidth(self):
        self.bandwidth = None

    @property
    def band_string(self) -> str:
        """A string representation of the band specified (e.g. "210 (4) nm")"""
        out = f'{float(self.wavelength)}'
        if self.bandwidth != 0:
            out += f' ({float(self.bandwidth)})'
        return out + f' {self.wavelength.unit}'

    def _update_unreferenced_intensities(self):
        """updates the unreferenced intensities attribute with the current band values"""
        if self.spectrum is not None and all([self.wavelength, self.bandwidth]):
            self._unreferenced_intensities = self.spectrum.get_intensities_from_signal(self)

    @property
    def reference(self) -> "DADSignal":
        """reference band for the signal band"""
        return self._reference

    @reference.setter
    def reference(self, value: Union[str, "DADSignal", "DADSignalInfo"]):
        if type(value) is str:
            parsed = self.get_values_from_agilent_string(value)
            value = DADSignal(
                **parsed,
                spectrum=self.spectrum,
            )
        elif isinstance(value, DADSignalInfo):
            value = self.create_from_DADSignalInfo(value, self.spectrum)
        self._reference = value

    @reference.deleter
    def reference(self):
        self._reference = None

    @property
    def unreferenced_intensities(self) -> np.ndarray:
        """unreferenced intensities for the band"""
        return self._unreferenced_intensities

    @property
    def mean_unreferenced_intensities(self) -> np.ndarray:
        """mean unreferenced intensities for the band"""
        return self.unreferenced_intensities.mean(axis=1)

    @property
    def mean_referenced_intensities(self) -> np.ndarray:
        """mean referenced band (mean unreferenced intensities minus the mean intensities of the reference)"""
        if self.reference is not None:
            return self.mean_unreferenced_intensities - self.reference.mean_unreferenced_intensities
        else:
            return self.mean_unreferenced_intensities

    @property
    def retention_times(self) -> np.ndarray:
        """retention times associated with the intensity array"""
        return self.spectrum.retention_times

    @classmethod
    def create_from_DADSignalInfo(cls, obj: DADSignalInfo, spectrum: DADSpectrum) -> "DADSignal":
        """generates a DADSignal object from a DADSignalInfo object and a spectrum"""
        return cls(
            wavelength=obj.wavelength,
            bandwidth=obj.bandwidth,
            reference=obj.reference,
            name=obj.name,
            spectrum=spectrum,
        )

    def as_iterable_data_table(self):
        """
        Returns an iterable which yields a data table with  appropriate headers and data

        :return: data table as iterable
        """
        if self.spectrum is None:
            raise AttributeError(f'no spectral data is associated with the instance')
        if self.reference is not None:  # if the signal is referenced
            # yield header row
            yield (
                f'Retention Time ({self.DEFAULT_TIME_UNIT})',
                f'{self.wavelength} intensity ({self.spectrum.data.yunits})',
                f'{self.reference.wavelength} reference intensity ({self.spectrum.data.yunits})',
                f'{self.wavelength} referenced intensity ({self.spectrum.data.yunits})',
            )
            for rt, intensity, ref_intensity, referenced in zip(
                self.retention_times,
                self.mean_unreferenced_intensities,
                self.reference.mean_unreferenced_intensities,
                self.mean_referenced_intensities,
            ):
                yield (rt, intensity, ref_intensity, referenced)
        else:  # if not referenced
            # yield header row
            yield (
                f'Retention Time ({self.DEFAULT_TIME_UNIT})',
                f'{self.wavelength} intensity ({self.spectrum.data.yunits})',
            )
            for rt, intensity in zip(
                    self.retention_times,
                    self.mean_unreferenced_intensities,
            ):
                yield (rt, intensity)

    def as_data_table(self) -> list:
        """
        Returns the signal as a list-style data table with appropriate headers and data

        :return: data table
        """
        return [list(line) for line in self.as_iterable_data_table()]

    def write_signal_to_csv(self, filename: str, overwrite: bool = False) -> str:
        """
        Writes the signal intensities to the specified csv file.

        :param filename: file name to write to
        :param overwrite: whether to overwrite the file if it already exists
        :return: file path that was written
        """
        # if the file already exists and overwrite was not specified, return
        if os.path.isfile(filename) and overwrite is False:
            return
        data = "\n".join(
            ",".join([str(val) for val in line])
            for line in self.as_iterable_data_table()
        )
        with open(filename, 'wt') as f:
            f.write(data)
        return filename


class HPLCSample(HPLCSampleInfo):
    def __init__(self,
                 sample_file_name: str,
                 method_name: str,
                 signals: Union[List[DADSignalInfo], List[DADSignal], List[str]],
                 datetimestamp: Union[str, datetime.datetime] = None,
                 dad_spectrum: DADSpectrum = None,
                 ms_spectrum: MSSpectrum = None,
                 directory: str = None,
                 ):
        """
        Data class for describing an HPLC sample containing metadata and spectral data.

        :param sample_file_name: name for sample
        :param datetimestamp: date and time stamp for when the sample was run
        :param method_name: name of method used to run the sample
        :param signals: list of signals associated with the run
        :param dad_spectrum: DADSpectrum object with loaded data
        :param directory: directory path where the sample may be found
        """
        HPLCSampleInfo.__init__(
            self,
            sample_file_name=sample_file_name,
            method_name=method_name,
            signals=signals,
            datetimestamp=datetimestamp
        )
        self.spectrum: DADSpectrum = dad_spectrum
        self.ms_spectrum: MSSpectrum = ms_spectrum
        self.signals = [DADSignal.create_from_DADSignalInfo(signal, self.spectrum) for signal in self.signals]
        self.directory = directory

    @classmethod
    def create_from_D_file(cls, file_path: Union[pathlib.Path, str]) -> "HPLCSample":
        """
        Creates an HPLCSample instance from a .D file.

        :param file_path: file path to Agilent .D folder
        :return: instantiated HPLCSample with loaded data
        """
        if isinstance(file_path, pathlib.Path) is False:
            file_path = pathlib.Path(file_path)
        if file_path.suffix != '.D' is False or os.path.isdir(file_path) is False:
            raise ValueError(f'The file path "{file_path}" does not appear to be a valid Agilent sample directory')
        sample_info_values = HPLCSampleInfo.find_and_get_metadata(file_path)
        try:
            dad_spectrum = DADSpectrum.create_from_D_file(file_path)
        except FileNotFoundError:
            dad_spectrum = None
        try:
            ms_spectrum = MSSpectrum.create_from_D_file(file_path)
        except FileNotFoundError:
            ms_spectrum = None
        return cls(
            **sample_info_values,
            dad_spectrum=dad_spectrum,
            ms_spectrum=ms_spectrum,
            directory=file_path,
        )

    @classmethod
    def create_from_acaml(cls, acaml: Union[str, xml.etree.ElementTree.ElementTree]) -> "HPLCSampleInfo":
        """not supported for HPLCSample class"""
        raise NotImplementedError(f'Creation of {cls.__class__.__name__} from acaml is not supported, use '
                                  f'create_from_D_file.')

    @classmethod
    def create_from_xml(cls, xml_path: Union[str, xml.etree.ElementTree.ElementTree]) -> "HPLCSampleInfo":
        """
        Creates sample structure from a Sample.xml file (old style metadata) in the desired .D folder)

        :param xml_path: path to xml file or parsed element tree root
        :return: parsed Sample instance
        """
        raise NotImplementedError(f'Creation of {cls.__class__.__name__} from xml is not supported, use '
                                  f'create_from_D_file.')

    def add_signal(self, new_signal: Union[DADSignalInfo, dict, str]) -> DADSignal:
        """
        Adds a new signal to the HPLCSample instance.

        :param new_signal: new signal to add. Supported inputs are Agilent specification strings
            (e.g. 'DAD1 A, Sig=210,4 Ref=360,100')  DADSignalInfo objects or a dictionary of keyword arguments for
            instantiating the same.
        :return: the created signal
        """
        if type(new_signal) is str:
            # assumed to be agilent specification string
            new_signal = DADSignalInfo.create_from_agilent_string(new_signal)
        # convert to DADSignal
        if isinstance(new_signal, DADSignalInfo):
            new_signal = DADSignal.create_from_DADSignalInfo(
                new_signal,
                spectrum=self.spectrum,
            )
        elif type(new_signal) is dict:
            new_signal = DADSignal(**new_signal, spectrum=self.spectrum)
        self.signals.append(new_signal)
        return new_signal

    def write_signals_to_csv(self,
                             directory: Union[str, pathlib.Path] = None,
                             overwrite: bool = False
                             ) -> List[str]:
        """
        Writes the signals to csv in the directory specified. If no directory is specified, the csv files will be
        written to the directory path specified in the directory attribute of the instance.

        :param directory: directory path
        :param overwrite: whether to overwrite files if they already exist
        :return: file paths written
        """
        if directory is None:
            if self.directory is None:
                raise ValueError('No directory was specified and the instance does not have a directory attribute.')
            directory = self.directory
        directory = pathlib.Path(directory).absolute()
        written_paths = []
        for signal in self.signals:
            file_name = directory / f'{signal}.csv'
            signal.write_signal_to_csv(
                file_name,
                overwrite=overwrite,
            )
            if file_name is not None:  # if a file was written
                written_paths.append(str(file_name))
        # write TAC
        tac_file_name = directory / 'Total Absorbance Chromatogram.csv'
        written_paths.append(str(tac_file_name))
        tac_data = "\n".join(
            f"{rt},{intensity}"
            for rt, intensity in zip(
                self.spectrum.retention_times,
                self.spectrum.total_absorbance_chromatogram,
            )
        )
        with open(tac_file_name, 'wt') as f:
            f.write(
                f'Retention Time (min),'
                f'Total Absorbance Chromatogram ({self.spectrum.chromatogram.yunits})\n'
            )
            f.write(tac_data)
        return written_paths

    def write_signals_to_xlsx(self, output_file: Union[str, pathlib.Path] = None) -> str:
        """
        Writes the signals to a single excel file.

        :param output_file: target file path. If this is not specified
        :return: path to the written file
        """
        if output_file is None:
            try:
                sample_name = self.spectrum.info["name"]
            except UnicodeDecodeError:
                sample_name = 'no Sample Name'
            output_file = pathlib.Path(self.spectrum.filename).parent / f'{sample_name}.xlsx'
        elif isinstance(output_file, pathlib.Path) is False:
            output_file = pathlib.Path(output_file)

        if output_file.is_absolute() is False:
            output_file = output_file.absolute()

        excel = openpyxl.Workbook()

        for signal in self.signals:
            # create sheet
            excel.create_sheet(signal.agilent_specification_string)
            sheet = excel[signal.agilent_specification_string]

            # todo add intensity units
            # save signals to sheets
            if signal.reference is not None:
                headers = [
                    f'Retention Time ({signal.DEFAULT_TIME_UNIT})',  # retention time
                    f'{signal.band_string}',  # unreferenced values
                    f'Reference {signal.reference.band_string}',  # reference values
                    'Referenced values',  # referenced values
                ]
                iterable = zip(
                    signal.retention_times,
                    signal.mean_unreferenced_intensities,
                    signal.reference.mean_unreferenced_intensities,
                    signal.mean_referenced_intensities,
                )
            else:
                headers = [
                    f'Retention Time ({signal.DEFAULT_TIME_UNIT})',  # retention time
                    f'{signal.band_string}',  # unreferenced values
                ]
                iterable = zip(
                    signal.retention_times,
                    signal.mean_unreferenced_intensities,
                )

            sheet.append(headers)
            for vals in iterable:
                sheet.append(vals)

        # append TAC
        tac_key = 'Total Absorbance Chromatogram'
        excel.create_sheet(tac_key)
        sheet = excel[tac_key]
        sheet.append([
            f'Retention Time (min)',  # todo retrieve from object
            f'Intensity ({self.spectrum.chromatogram.yunits})',
        ])
        for time, inten in zip(self.spectrum.retention_times, self.spectrum.total_absorbance_chromatogram):
            sheet.append([time, inten])

        if 'Sheet' in excel:
            excel.remove(excel['Sheet'])

        excel.save(output_file)
        return output_file
