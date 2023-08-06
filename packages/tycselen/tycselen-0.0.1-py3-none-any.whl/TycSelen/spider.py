#coding=utf-8

import os
import time
import re
import xlsxwriter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import numbers, Alignment
from openpyxl import load_workbook
from pymongo import MongoClient
import pandas as pd
import numpy as np

#带入程序包
import json
import pandas as pd
import jieba
import jieba.posseg as psg
import collections
from wordcloud import WordCloud, ImageColorGenerator, STOPWORDS
import matplotlib.pyplot as plt
from cnsenti import Sentiment
from PIL import Image


rootPath=r'./'

class Excel():
    #初始化
    def __init__(self):
        self.path = ''
    # 创建或打开Excel
    def open(self, path, flag=False):
        self.path = path.encode('utf-8').decode('utf-8')
        if os.path.exists(self.path) and flag:
            self.wb=load_workbook(self.path)
        else:
            self.wb=Workbook()
        self.sheet_init('Sheet')
        # 初始化工作表
    def sheet_init(self, name):
        self.ws = self.wb[name] if name in self.wb.sheetnames else self.wb.create_sheet(name.encode('utf-8').decode('utf-8'))
        # 设置列宽
        # self.ws.column_dimensions.fitToWidth=False
        # self.ws.column_dimensions['A'].width = 10
        # self.ws.column_dimensions['B'].width = 30
        # self.ws.column_dimensions['C'].width = 28
        # self.ws.column_dimensions['D'].width = 30
        # self.ws.column_dimensions['L'].width = 3000

    # 按行写入Excel
    def write_row(self, row):
        self.ws.append(row)
    # 关闭Excel
    def close(self):
        self.wb.close()
    # 格式化输出
    def format(self):
        for row in self.ws.rows:
            for i in range(self.ws.min_column, self.ws.max_column):
                row[i].alignment = Alignment(horizontal='right')
                # row[i].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1 if i<0 else '#,###'
    # 保存并关闭Excel
    def save(self):
        self.wb.save(filename=self.path)
        # self.wb.close()


class Mongo():
    #初始化
    def __init__(self,stamp=None):
        self.excel = Excel()  # Excel对象
        self.stamp = (datetime.now()).strftime('%Y%m%d') if stamp is None else stamp
        self.client = MongoClient("localhost", 27017)
        self.db=self.client['tianyancha']
        self.db2 = self.client['tianyancha']

    def getStamp(self):
        return self.utc2local(self.local2utc(datetime.now())).strftime('%Y%m%d')

    # excel汇总去重
    def huizong(self, dirpath):
        df_out = None
        self._matches = self.db2['df_qichacha']
        print('Waiting...')
        print(len(os.listdir(dirpath)))
        index=0
        for filename in os.listdir(dirpath):
            index+=1
            filepath = os.path.join(dirpath, filename).replace('\\', '/')
            print(index, filepath)
            # data = pd.read_excel(filepath, header=0, skiprows=1, nrows=1)
            df_in = pd.read_excel(filepath, header=1)
            print(df_in)
            df_value = json.loads(df_in.T.to_json()).values()
            self._matches.insert_many(df_value)

        print('去重')
        # df_out.drop_duplicates(subset=['mid'], keep='first', inplace=True)
        # df_out.to_excel('search_汇总.xlsx', engine='xlsxwriter')

    # 下载数据并保存
    def export(self):
        # 打开Excel
        # # 大数据，区块链，其他，人工智能，网络金融，云计算
        self.excel.open(rootPath + 'ExcelData/search_汇总1.xlsx')
        title='company,class, keyword,title,url,summary,source,stamp'.split(',')
        self.excel.write_row(title)
        # 写入数据
        self._matches = self.db['search_汇总1']
        # self._matches = self.db['guyulab_2']
        # self._matches = self.db['zhenshigushi1_2']
        try:
            matches = self._matches.find()
        except:
            print (self.stamp + " error, quit()")
            quit()
        for match in matches:
            try:
                # arctile=match['html']
                # arctile='error' if (arctile is None) or (len(arctile)<2000) else str(arctile)
                body=[match['company'],match['fenlei'],match['keyword'],match['title'],match['url'],match['summary'],match['source'],match['stamp']]
                self.excel.write_row(body)
            except:
                continue
        # 关闭并保存Excel
        # self.excel.format()
        self.excel.save()

    def pb_import(self):
        self._matches = self.db2['df_qichaccha']
        df=pd.read_excel(r'search_汇总3.xlsx')
        # print(json.loads(df.T.to_json()))
        df_value=json.loads(df.T.to_json()).values()
        # print(df_value)
        self._matches.insert_many(df_value)


    # 下载数据并保存
    def pb_export(self):
        # 大数据，区块链，其他，人工智能，网络金融，云计算
        self._matches = self.db['search_qichacha45000']
        self._matches2 = self.db['search_qichacha_zero']
        try:
            # 不为空查询
            matches = self._matches.find({'匹配名称': {'$in': ['NaN']}})
            matches2 = self._matches2.find()
            symbol_list=[x['symbol'] for x in matches2]
            # matches = self._matches.find({'datetime': {'$lte':1598528644,'$gte':1567424487}, 'content': {'$not': {'$in': [None]}}, 'author': {'$not': {'$in': ['']}}})
            # matches = self._matches.find(
            #     {'datetime': {'$lte': 1599015994, '$gte': 1567135988}, 'content': {'$in': [None]}})

            print(matches.count())

            for match in matches:
                try:
                    if match['symbol'] not in symbol_list:
                        if len(match['查询结果'])==0:
                            # match['匹配名称']=match['查询结果'][0][0]
                            # match['网址'] = match['查询结果'][0][1]
                            # print(type(match['查询结果']))
                            self._matches2.insert_one(match)
                    # company=match['company']
                    # # fenlei = match['fenlei']
                    # keyword = match['keyword']
                    # title = match['title']
                    # summary = match['summary']
                    #
                    # content=title+summary
                    # flag=(company in content) and (keyword in content)
                    # print(flag)
                    # if flag:
                    #     self._matches2.insert_one(match)

                except:
                    continue
        except:
            print (self.stamp + " error, quit()")
            quit()

    # df求交集，差集，出结果
    def pb_export2(self):
        try:
            table_name = 'df_qichacha'
            self._matches = self.db2[table_name]
            matches = self._matches.find()
            df1 = pd.DataFrame(list(matches))
            del df1['_id']
            df1 = df1[['企业名称', '成立日期', '所属省份', '纳税人识别号', '企业类型', '所属行业', '企业地址']]
            # df1 = pd.read_excel('input3.xlsx')
            # df1 = df1[['企业名称']]
            print(df1)
            df2=pd.read_excel('匹配2.xlsx')
            df2 = df2[['公司名称', '匹配名称']]
            print(df2)
            # 合集
            df=pd.merge(df2, df1, left_on='匹配名称', right_on='企业名称', how='left')
            df.drop_duplicates(subset=['公司名称'], keep='first', inplace=True)
            print(df)
            # 交集（完全匹配的）
            df = df.loc[df['企业名称'].str.len().gt(0)]
            # 差集（未完全匹配的）
            # df = df.loc[df['企业名称'].isnull()]


            df.reset_index(drop=True, inplace=True)
            print(df)
            df.to_excel('合并2.xlsx')

        except Exception as e:
            print (e)

    # 预处理待查询公司名单
    def pb_export3(self):
        try:
            df=pd.read_excel('input2.xlsx')
            # print(df)
            # df['公司名称'] = df[df['公司名称'].apply(str)]
            # df = df.dropna(how='any', subset=['公司名称'])
            # df = df.loc[df[['公司名称']].notna().all(1)]
            # df=df[df['公司名称'].str.replace('(', '（').replace(')', '）')]
            for index, row in df.iterrows():
                print(row)
                try:
                    # print(df.iloc[index, 0])
                    df.iloc[index, 0] = row[0].replace('(', '（').replace(')', '）')
                except Exception as e:
                    print(e)
            # df.drop_duplicates(subset=['企业名称'], keep='first', inplace=True)
            df.to_excel('input3.xlsx')

        except Exception as e:
            print(e)

    # df求交集，差集，出结果
    def pb_export4(self):
        try:
            df1 = pd.read_excel('模糊匹配.xlsx')
            df1 = df1[['公司名称', '企业名称']]
            print(df1)
            df2=pd.read_excel('匹配2.xlsx')
            df2 = df2['公司名称']
            print(df2)
            # 合集
            df=pd.merge(df1, df2, left_on='公司名称', right_on='公司名称', how='right')
            df.drop_duplicates(subset=['公司名称'], keep='first', inplace=True)
            print(df)
            # 交集（完全匹配的）
            # df = df.loc[df['企业名称'].str.len().gt(0)]
            # 差集（未完全匹配的）
            df = df.loc[df['企业名称'].isnull()]


            # df.reset_index(drop=True, inplace=True)
            print(df)
            df.to_excel('合并2.xlsx')

        except Exception as e:
            print (e)

    # search求交集，差集，出待查名单
    def pb_export5(self):
        try:
            table_name = 'search_qichacha'
            self._matches = self.db2[table_name]
            # matches = self._matches.find()
            matches = self._matches.find({'匹配状态': 1})
            df1 = pd.DataFrame(list(matches))
            # del df1['_id']
            df1=df1[['symbol', '匹配名称']]
            print(df1)
            df2=pd.read_excel('待查询.xlsx')
            df2 = df2['公司名称']
            print(df2)
            # 合集
            df=pd.merge(df2, df1, left_on='公司名称', right_on='symbol', how='left')
            df.drop_duplicates(subset=['symbol'], keep='first', inplace=True)
            print(df)
            # 交集（完全匹配的）
            df = df.loc[df['symbol'].str.len().gt(0)]
            # 差集（未完全匹配的）
            # df = df.loc[df['企业名称'].isnull()]


            df.reset_index(drop=True, inplace=True)
            print(df)
            df.to_excel('合并2.xlsx')

        except Exception as e:
            print (e)
            quit()



    # 下载数据并保存
    def pb_export6(self):

        try:
            df1=pd.read_excel('search_qichacha_mohu.xlsx')
            print(df1)
            df2=pd.read_excel('qichacha_模糊匹配.xlsx')
            print(df2)

            df=pd.merge(df1, df2, left_on='匹配名称', right_on='企业名称', how='inner')
            df.drop_duplicates(subset=['公司名称'], keep='first', inplace=True)
            print(df)
            df.to_excel('合并.xlsx')

        except:
            print (self.stamp + " error, quit()")
            quit()

    # 下载数据并保存
    def pb_export7(self):
        try:
            table_name = 'df_qichacha'
            self._matches1 = self.db2[table_name]
            matches1 = self._matches1.find()
            df1 = pd.DataFrame(list(matches1))
            del df1['_id']
            print(df1)
            df2=pd.read_excel('search_qichacha_clear.xlsx')
            print(df2)

            df=pd.merge(df2, df1, left_on='匹配名称', right_on='企业名称', how='inner')
            df.drop_duplicates(subset=['公司名称'], keep='first', inplace=True)
            print(df)
            df.to_excel('合并2.xlsx')

        except Exception as e:
            print (e)
            quit()

# 主程序入口
if __name__ == '__main__':
    # Ctrl+ / 切换注释
    # db = Mongo("2019-04-03")   #下载历史数据（但必须已写入数据库）

    # 显示所有列
    # pd.set_option('display.max_columns', None)
    # # 显示所有行
    # pd.set_option('display.max_rows', None)
    # # 设置value的显示长度为100，默认为50
    # pd.set_option('max_colwidth', 100)
    # plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文标签
    # plt.rcParams['axes.unicode_minus'] = False



    db = Mongo()
    db.pb_export4()
    # db.pb_import()
    # db.export()
    # db.huizong(r'C:\Users\iseuwei\Desktop\a2012057\1.4\汇总')


    print('\nTS:'+datetime.now().strftime('%Y-%m-%d %H:%M:%S'))



